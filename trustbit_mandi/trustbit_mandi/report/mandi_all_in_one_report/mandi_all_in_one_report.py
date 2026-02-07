# Copyright (c) 2026, Trustbit Software and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.utils import flt, getdate


def execute(filters=None):
	columns = get_columns()
	data = get_data(filters)
	report_summary = get_report_summary(data)
	return columns, data, None, None, report_summary


def get_columns():
	return [
		{"fieldname": "sno", "label": _("S.No."), "fieldtype": "Int", "width": 50},
		{"fieldname": "contract_date", "label": _("Contract Date"), "fieldtype": "Date", "width": 100},
		{"fieldname": "contract_number", "label": _("Contract No."), "fieldtype": "Data", "width": 110},
		{"fieldname": "as_flag", "label": _("A/S"), "fieldtype": "Data", "width": 50},
		{"fieldname": "farmer_name", "label": _("Farmer Name"), "fieldtype": "Data", "width": 130},
		{"fieldname": "address", "label": _("Address"), "fieldtype": "Data", "width": 100},
		{"fieldname": "phone_number", "label": _("Phone"), "fieldtype": "Data", "width": 100},
		{"fieldname": "gsm", "label": _("Commodity"), "fieldtype": "Data", "width": 90},
		{"fieldname": "transaction_no", "label": _("Transaction No."), "fieldtype": "Data", "width": 120},
		{"fieldname": "expected_bag", "label": _("Exp. Bag"), "fieldtype": "Float", "width": 70},
		{"fieldname": "actual_bag", "label": _("Actual Bag"), "fieldtype": "Float", "width": 80},
		{"fieldname": "actual_weight", "label": _("Weight (Qtl)"), "fieldtype": "Float", "width": 90, "precision": 2},
		{"fieldname": "auction_rate", "label": _("Rate"), "fieldtype": "Currency", "width": 80},
		{"fieldname": "amount", "label": _("Amount"), "fieldtype": "Currency", "width": 100},
		{"fieldname": "rounded_amount", "label": _("Rounded Amt"), "fieldtype": "Currency", "width": 100},
		{"fieldname": "hamali", "label": _("Hamali"), "fieldtype": "Currency", "width": 80},
		{"fieldname": "net_amount", "label": _("Net Amount"), "fieldtype": "Currency", "width": 110},
		{"fieldname": "mandi_tax", "label": _("Mandi Tax"), "fieldtype": "Currency", "width": 90},
		{"fieldname": "nirashrit_tax", "label": _("Nirashrit Tax"), "fieldtype": "Currency", "width": 90},
		{"fieldname": "total_tax", "label": _("Total Tax"), "fieldtype": "Currency", "width": 90},
		{"fieldname": "payment_status", "label": _("Pay Status"), "fieldtype": "Data", "width": 90},
		{"fieldname": "pay_date", "label": _("Payment Date"), "fieldtype": "Date", "width": 100},
		{"fieldname": "payment_mode", "label": _("Pay Mode"), "fieldtype": "Data", "width": 90},
		{"fieldname": "bank_name", "label": _("Bank"), "fieldtype": "Data", "width": 100},
		{"fieldname": "account_number", "label": _("Account No."), "fieldtype": "Data", "width": 120},
		{"fieldname": "ifsc_code", "label": _("IFSC"), "fieldtype": "Data", "width": 100},
		{"fieldname": "name", "label": _("ID"), "fieldtype": "Link", "options": "Grain Purchase", "width": 120},
	]


def get_data(filters):
	from_date = filters.get("from_date")
	to_date = filters.get("to_date")
	as_flag = filters.get("as_flag")
	payment_status = filters.get("payment_status")
	payment_mode = filters.get("payment_mode")
	gsm_filter = filters.get("gsm")
	farmer_name = filters.get("farmer_name")

	today = frappe.utils.nowdate()

	if not from_date:
		from_date = frappe.utils.add_months(today, -1)
	if not to_date:
		to_date = today

	conditions = ["docstatus < 2", "contract_date >= %s", "contract_date <= %s"]
	values = [from_date, to_date]

	if as_flag:
		conditions.append("as_flag = %s")
		values.append(as_flag)

	if payment_status:
		conditions.append("payment_status = %s")
		values.append(payment_status)

	if payment_mode:
		conditions.append("payment_mode = %s")
		values.append(payment_mode)

	if gsm_filter:
		conditions.append("gsm = %s")
		values.append(gsm_filter)

	if farmer_name:
		conditions.append("farmer_name LIKE %s")
		values.append("%" + farmer_name + "%")

	sql = """
		SELECT
			name,
			contract_date,
			contract_number,
			as_flag,
			farmer_name,
			address,
			phone_number,
			gsm,
			transaction_no,
			expected_bag,
			actual_bag,
			actual_weight,
			auction_rate,
			amount,
			rounded_amount,
			hamali,
			net_amount,
			mandi_tax,
			nirashrit_tax,
			total_tax,
			payment_status,
			pay_date,
			payment_mode,
			bank_name,
			account_number,
			ifsc_code
		FROM `tabGrain Purchase`
		WHERE {conditions}
		ORDER BY contract_date ASC, farmer_name ASC
	""".format(conditions=" AND ".join(conditions))

	raw_data = frappe.db.sql(sql, values, as_dict=True)

	# Add serial numbers
	sno = 0
	for row in raw_data:
		sno += 1
		row["sno"] = sno

	# Add total row
	if raw_data:
		total_weight = sum([(r.get("actual_weight") or 0) for r in raw_data])
		total_amount = sum([(r.get("amount") or 0) for r in raw_data])
		total_rounded = sum([(r.get("rounded_amount") or 0) for r in raw_data])
		total_hamali = sum([(r.get("hamali") or 0) for r in raw_data])
		total_net = sum([(r.get("net_amount") or 0) for r in raw_data])
		total_mandi_tax = sum([(r.get("mandi_tax") or 0) for r in raw_data])
		total_nirashrit_tax = sum([(r.get("nirashrit_tax") or 0) for r in raw_data])
		total_total_tax = sum([(r.get("total_tax") or 0) for r in raw_data])
		total_exp_bag = sum([(r.get("expected_bag") or 0) for r in raw_data])
		total_act_bag = sum([(r.get("actual_bag") or 0) for r in raw_data])

		raw_data.append({
			"sno": "",
			"farmer_name": "<b>TOTAL</b>",
			"expected_bag": total_exp_bag,
			"actual_bag": total_act_bag,
			"actual_weight": total_weight,
			"amount": total_amount,
			"rounded_amount": total_rounded,
			"hamali": total_hamali,
			"net_amount": total_net,
			"mandi_tax": total_mandi_tax,
			"nirashrit_tax": total_nirashrit_tax,
			"total_tax": total_total_tax,
		})

	return raw_data


def get_report_summary(data):
	# Exclude the total row for summary calculation
	report_data = [r for r in data if r.get("sno") != ""]

	total_weight = sum([(r.get("actual_weight") or 0) for r in report_data])
	total_amount = sum([(r.get("amount") or 0) for r in report_data])
	total_hamali = sum([(r.get("hamali") or 0) for r in report_data])
	total_net = sum([(r.get("net_amount") or 0) for r in report_data])

	return [
		{"value": total_weight, "label": _("Total Weight (Qtl)"), "datatype": "Float", "indicator": "Blue"},
		{"value": total_amount, "label": _("Total Amount"), "datatype": "Currency", "indicator": "Blue"},
		{"value": total_hamali, "label": _("Total Hamali"), "datatype": "Currency", "indicator": "Orange"},
		{"value": total_net, "label": _("Total Net Amount"), "datatype": "Currency", "indicator": "Green"},
	]


def _get_payment_data(filters):
	"""Common helper to fetch grain purchase payment data for exports"""
	from_date = filters.get("from_date") or frappe.utils.add_months(frappe.utils.nowdate(), -1)
	to_date = filters.get("to_date") or frappe.utils.nowdate()
	as_flag = filters.get("as_flag")

	conditions = ["docstatus < 2", "contract_date >= %s", "contract_date <= %s"]
	values = [from_date, to_date]

	if as_flag:
		conditions.append("as_flag = %s")
		values.append(as_flag)

	payment_status = filters.get("payment_status")
	if payment_status:
		conditions.append("payment_status = %s")
		values.append(payment_status)

	payment_mode = filters.get("payment_mode")
	if payment_mode:
		conditions.append("payment_mode = %s")
		values.append(payment_mode)

	gsm_filter = filters.get("gsm")
	if gsm_filter:
		conditions.append("gsm = %s")
		values.append(gsm_filter)

	sql = """
		SELECT
			name, contract_date, contract_number, as_flag,
			farmer_name, address, phone_number, gsm,
			transaction_no, expected_bag, actual_bag,
			actual_weight, auction_rate, amount, rounded_amount,
			hamali, net_amount, mandi_tax, nirashrit_tax, total_tax,
			payment_status, pay_date, payment_mode,
			bank_name, account_number, ifsc_code, branch
		FROM `tabGrain Purchase`
		WHERE {conditions}
		ORDER BY contract_date ASC, farmer_name ASC
	""".format(conditions=" AND ".join(conditions))

	return frappe.db.sql(sql, values, as_dict=True)


@frappe.whitelist()
def export_neft_rtgs_excel(**filters):
	"""Export NEFT/RTGS Excel format for bulk bank payment"""
	from io import BytesIO
	import openpyxl
	from openpyxl.styles import Font, Alignment, Border, Side

	data = _get_payment_data(filters)
	if not data:
		frappe.throw(_("No data found for the selected filters"))

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "NEFT_RTGS Payment"

	# Header style
	header_font = Font(bold=True, size=11)
	border = Border(
		left=Side(style='thin'), right=Side(style='thin'),
		top=Side(style='thin'), bottom=Side(style='thin')
	)

	headers = [
		"S.No.", "Payment Date", "Beneficiary Name", "Account Number",
		"IFSC Code", "Bank Name", "Branch", "Amount (Rs.)",
		"Payment Mode", "Contract No.", "Commodity", "Phone",
		"Remarks"
	]

	for col, header in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col, value=header)
		cell.font = header_font
		cell.alignment = Alignment(horizontal='center')
		cell.border = border

	for idx, row in enumerate(data, 1):
		values = [
			idx,
			str(row.get("pay_date") or row.get("contract_date") or ""),
			row.get("farmer_name") or "",
			row.get("account_number") or "",
			row.get("ifsc_code") or "",
			row.get("bank_name") or "",
			row.get("branch") or "",
			flt(row.get("net_amount")),
			row.get("payment_mode") or "NEFT",
			row.get("contract_number") or "",
			row.get("gsm") or "",
			row.get("phone_number") or "",
			"Grain Purchase Payment"
		]
		for col, val in enumerate(values, 1):
			cell = ws.cell(row=idx + 1, column=col, value=val)
			cell.border = border

	# Auto-width columns
	for col in ws.columns:
		max_length = 0
		for cell in col:
			if cell.value:
				max_length = max(max_length, len(str(cell.value)))
		ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 30)

	output = BytesIO()
	wb.save(output)
	output.seek(0)

	frappe.response['filename'] = 'NEFT_RTGS_Payment_{}.xlsx'.format(
		frappe.utils.nowdate()
	)
	frappe.response['filecontent'] = output.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def export_dtr_excel(**filters):
	"""Export DTR (Direct Transfer) Excel format for bank"""
	from io import BytesIO
	import openpyxl
	from openpyxl.styles import Font, Alignment, Border, Side

	data = _get_payment_data(filters)
	if not data:
		frappe.throw(_("No data found for the selected filters"))

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "DTR Payment"

	header_font = Font(bold=True, size=11)
	border = Border(
		left=Side(style='thin'), right=Side(style='thin'),
		top=Side(style='thin'), bottom=Side(style='thin')
	)

	headers = [
		"S.No.", "Transaction Date", "Beneficiary Name",
		"Beneficiary Account No.", "IFSC Code", "Bank Name",
		"Transfer Amount (Rs.)", "Transfer Type", "Narration",
		"Contract No.", "Phone No."
	]

	for col, header in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col, value=header)
		cell.font = header_font
		cell.alignment = Alignment(horizontal='center')
		cell.border = border

	for idx, row in enumerate(data, 1):
		values = [
			idx,
			str(row.get("pay_date") or row.get("contract_date") or ""),
			row.get("farmer_name") or "",
			row.get("account_number") or "",
			row.get("ifsc_code") or "",
			row.get("bank_name") or "",
			flt(row.get("net_amount")),
			"DTR",
			"Mandi Grain Purchase - {}".format(row.get("contract_number") or ""),
			row.get("contract_number") or "",
			row.get("phone_number") or "",
		]
		for col, val in enumerate(values, 1):
			cell = ws.cell(row=idx + 1, column=col, value=val)
			cell.border = border

	for col in ws.columns:
		max_length = 0
		for cell in col:
			if cell.value:
				max_length = max(max_length, len(str(cell.value)))
		ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 30)

	output = BytesIO()
	wb.save(output)
	output.seek(0)

	frappe.response['filename'] = 'DTR_Payment_{}.xlsx'.format(
		frappe.utils.nowdate()
	)
	frappe.response['filecontent'] = output.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def export_krishi_upaj_excel(**filters):
	"""Export Krishi Upaj Mandi format - separate A and S Excel files"""
	from io import BytesIO
	import openpyxl
	from openpyxl.styles import Font, Alignment, Border, Side

	as_flag = filters.get("as_flag")
	if not as_flag:
		frappe.throw(_("Please select A or S flag for Krishi Upaj Mandi export"))

	data = _get_payment_data(filters)
	if not data:
		frappe.throw(_("No data found for A/S flag: {0}").format(as_flag))

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Krishi Upaj Mandi - {}".format(as_flag)

	header_font = Font(bold=True, size=11)
	border = Border(
		left=Side(style='thin'), right=Side(style='thin'),
		top=Side(style='thin'), bottom=Side(style='thin')
	)

	# Krishi Upaj Mandi portal format columns
	headers = [
		"S.No.", "A/S", "Contract Date", "Contract No.",
		"Farmer Name", "Address", "Phone",
		"Commodity", "Expected Bag", "Actual Bag",
		"Weight (Qtl)", "Auction Rate",
		"Amount (Rs.)", "Hamali (Rs.)", "Net Amount (Rs.)",
		"Mandi Tax (Rs.)", "Nirashrit Tax (Rs.)", "Total Tax (Rs.)",
		"Transaction No."
	]

	for col, header in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col, value=header)
		cell.font = header_font
		cell.alignment = Alignment(horizontal='center', wrap_text=True)
		cell.border = border

	total_weight = 0
	total_amount = 0
	total_hamali = 0
	total_net = 0
	total_mandi_tax = 0
	total_nirashrit_tax = 0
	total_total_tax = 0

	for idx, row in enumerate(data, 1):
		total_weight += flt(row.get("actual_weight"))
		total_amount += flt(row.get("amount"))
		total_hamali += flt(row.get("hamali"))
		total_net += flt(row.get("net_amount"))
		total_mandi_tax += flt(row.get("mandi_tax"))
		total_nirashrit_tax += flt(row.get("nirashrit_tax"))
		total_total_tax += flt(row.get("total_tax"))

		values = [
			idx,
			row.get("as_flag") or as_flag,
			str(row.get("contract_date") or ""),
			row.get("contract_number") or "",
			row.get("farmer_name") or "",
			row.get("address") or "",
			row.get("phone_number") or "",
			row.get("gsm") or "",
			flt(row.get("expected_bag")),
			flt(row.get("actual_bag")),
			flt(row.get("actual_weight")),
			flt(row.get("auction_rate")),
			flt(row.get("amount")),
			flt(row.get("hamali")),
			flt(row.get("net_amount")),
			flt(row.get("mandi_tax")),
			flt(row.get("nirashrit_tax")),
			flt(row.get("total_tax")),
			row.get("transaction_no") or "",
		]
		for col, val in enumerate(values, 1):
			cell = ws.cell(row=idx + 1, column=col, value=val)
			cell.border = border

	# Total row
	total_row = len(data) + 2
	total_font = Font(bold=True, size=11)
	ws.cell(row=total_row, column=1, value="").border = border
	total_cell = ws.cell(row=total_row, column=5, value="TOTAL")
	total_cell.font = total_font
	total_cell.border = border
	for col in [2, 3, 4, 6, 7, 8]:
		ws.cell(row=total_row, column=col, value="").border = border
	ws.cell(row=total_row, column=11, value=total_weight).font = total_font
	ws.cell(row=total_row, column=11).border = border
	ws.cell(row=total_row, column=13, value=total_amount).font = total_font
	ws.cell(row=total_row, column=13).border = border
	ws.cell(row=total_row, column=14, value=total_hamali).font = total_font
	ws.cell(row=total_row, column=14).border = border
	ws.cell(row=total_row, column=15, value=total_net).font = total_font
	ws.cell(row=total_row, column=15).border = border
	ws.cell(row=total_row, column=16, value=total_mandi_tax).font = total_font
	ws.cell(row=total_row, column=16).border = border
	ws.cell(row=total_row, column=17, value=total_nirashrit_tax).font = total_font
	ws.cell(row=total_row, column=17).border = border
	ws.cell(row=total_row, column=18, value=total_total_tax).font = total_font
	ws.cell(row=total_row, column=18).border = border

	for col in ws.columns:
		max_length = 0
		for cell in col:
			if cell.value:
				max_length = max(max_length, len(str(cell.value)))
		ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 30)

	output = BytesIO()
	wb.save(output)
	output.seek(0)

	frappe.response['filename'] = 'Krishi_Upaj_Mandi_{flag}_{date}.xlsx'.format(
		flag=as_flag, date=frappe.utils.nowdate()
	)
	frappe.response['filecontent'] = output.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def get_rtgs_entries(**filters):
	"""Get entries for bulk RTGS form printing"""
	data = _get_payment_data(filters)
	# Filter for RTGS/NEFT payment mode entries that have bank details
	rtgs_data = []
	for row in data:
		if row.get("account_number") and row.get("ifsc_code"):
			rtgs_data.append(row)
	return rtgs_data
